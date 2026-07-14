#!/usr/bin/env python3
"""
====================================================================
CAFE MENU TO AUTO-REPLY CONVERTER
====================================================================

This script converts your Excel menu file into auto-reply keywords
that you can upload to the WhatsApp Bot!

HOW TO USE:
1. Fill your menu in the Excel file (cafe_menu_template.xlsx)
2. Run this script
3. It will create keywords that you can add to the bot!

Author: Built for Indian Businesses 🇮🇳
====================================================================
"""

import json
import os
import sys
from pathlib import Path

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("⚠️ pandas not installed. Install with: pip install pandas openpyxl")

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl not installed. Install with: pip install openpyxl")


class CafeMenuConverter:
    """Convert Excel menu to auto-reply keywords"""
    
    def __init__(self):
        self.keywords = []
        self.cafe_name = "Our Cafe"
        self.address = "Your Address Here"
        self.phone = "Your Phone Number"
        
    def load_cafe_config(self):
        """Load cafe configuration"""
        config_path = Path("cafe_config.txt")
        if config_path.exists():
            with open(config_path, 'r') as f:
                lines = f.readlines()
                for line in lines:
                    if '=' in line:
                        key, value = line.strip().split('=', 1)
                        if key == 'cafe_name':
                            self.cafe_name = value
                        elif key == 'address':
                            self.address = value
                        elif key == 'phone':
                            self.phone = value
    
    def save_cafe_config(self, name, address, phone):
        """Save cafe configuration"""
        with open("cafe_config.txt", 'w') as f:
            f.write(f"cafe_name={name}\n")
            f.write(f"address={address}\n")
            f.write(f"phone={phone}\n")
        self.cafe_name = name
        self.address = address
        self.phone = phone
    
    def convert_from_excel(self, excel_path):
        """Convert Excel menu to keywords"""
        if not PANDAS_AVAILABLE:
            print("❌ Please install pandas: pip install pandas")
            return []
        
        try:
            # Read Excel file
            df = pd.read_excel(excel_path)
            
            # Print columns for debugging
            print(f"📋 Found columns: {list(df.columns)}")
            
            # Find the right columns
            keywords = []
            
            for index, row in df.iterrows():
                try:
                    # Try different column names
                    category = str(row.get('Category', row.get('category', row.get('Type', '')))
                    item_name = str(row.get('Item', row.get('item', row.get('Name', '')))
                    price = str(row.get('Price', row.get('price', row.get('Cost', '')))
                    keywords_list = str(row.get('Keywords', row.get('keywords', row.get('Search Terms', '')))
                    
                    # Skip empty rows
                    if item_name == 'nan' or item_name == '':
                        continue
                    
                    # Create keyword entry
                    item_keywords = []
                    
                    # Add item name words as keywords
                    for word in item_name.lower().split():
                        if len(word) > 2:
                            item_keywords.append(word)
                    
                    # Add custom keywords
                    if keywords_list != 'nan' and keywords_list:
                        for kw in keywords_list.split(','):
                            item_keywords.append(kw.strip().lower())
                    
                    # Create response
                    response = f"{item_name}"
                    if price != 'nan' and price:
                        response += f" - ₹{price}"
                    
                    keywords.append({
                        'category': category,
                        'keywords': item_keywords,
                        'response': response
                    })
                    
                except Exception as e:
                    print(f"⚠️ Error on row {index}: {e}")
                    continue
            
            self.keywords = keywords
            return keywords
            
        except Exception as e:
            print(f"❌ Error reading Excel: {e}")
            return []
    
    def convert_from_csv(self, csv_path):
        """Convert CSV menu to keywords"""
        if not PANDAS_AVAILABLE:
            print("❌ Please install pandas: pip install pandas")
            return []
        
        try:
            df = pd.read_csv(csv_path)
            return self.convert_from_dataframe(df)
        except Exception as e:
            print(f"❌ Error reading CSV: {e}")
            return []
    
    def convert_from_json(self, json_path):
        """Load keywords from JSON file"""
        try:
            with open(json_path, 'r') as f:
                data = json.load(f)
                return data.get('keywords', [])
        except Exception as e:
            print(f"❌ Error reading JSON: {e}")
            return []
    
    def generate_keywords_file(self, output_path="my_keywords.txt"):
        """Generate a keywords file for the bot"""
        lines = []
        
        # Add cafe info
        lines.append(f"# Cafe Keywords for {self.cafe_name}")
        lines.append(f"# Generated automatically")
        lines.append("")
        
        # Group by category
        categories = {}
        for kw in self.keywords:
            cat = kw.get('category', 'Other')
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(kw)
        
        for category, items in categories.items():
            lines.append(f"# === {category.upper()} ===")
            for item in items:
                for keyword in item.get('keywords', []):
                    response = item.get('response', '')
                    lines.append(f"{keyword}|{response}")
            lines.append("")
        
        # Write to file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        print(f"✅ Keywords saved to: {output_path}")
        return output_path
    
    def import_to_bot(self):
        """Import keywords directly to the bot database"""
        try:
            import sqlite3
            
            db_path = Path("data") / "whatsapp.db"
            db_path.parent.mkdir(exist_ok=True)
            
            conn = sqlite3.connect(str(db_path))
            c = conn.cursor()
            
            # Create table if not exists
            c.execute('''CREATE TABLE IF NOT EXISTS keywords (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                keyword TEXT UNIQUE,
                response TEXT
            )''')
            
            # Add keywords
            for kw in self.keywords:
                for keyword in kw.get('keywords', []):
                    response = kw.get('response', '')
                    try:
                        c.execute('INSERT INTO keywords (keyword, response) VALUES (?, ?)',
                                 (keyword.lower(), response))
                    except:
                        c.execute('UPDATE keywords SET response=? WHERE keyword=?',
                                 (response, keyword.lower()))
            
            conn.commit()
            conn.close()
            
            print(f"✅ Successfully imported {len(self.keywords)} items to bot database!")
            return True
            
        except Exception as e:
            print(f"❌ Error importing to bot: {e}")
            return False
    
    def show_preview(self, limit=10):
        """Show preview of keywords"""
        print(f"\n📋 Keyword Preview (first {limit}):\n")
        
        for i, kw in enumerate(self.keywords[:limit]):
            keywords_str = ', '.join(kw.get('keywords', [])[:3])
            response = kw.get('response', '')[:50]
            print(f"  {i+1}. Keywords: {keywords_str}")
            print(f"     Response: {response}...")
            print()
        
        if len(self.keywords) > limit:
            print(f"  ... and {len(self.keywords) - limit} more items")
        print()


def create_sample_excel():
    """Create a sample Excel menu file"""
    if not OPENPYXL_AVAILABLE:
        print("❌ Please install openpyxl: pip install openpyxl")
        return
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Menu"
    
    # Headers
    headers = ['Category', 'Item', 'Price', 'Keywords']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # Sample data
    sample_data = [
        ['Coffee', 'Espresso', '99', 'espresso,coffee,espresso shot'],
        ['Coffee', 'Cappuccino', '129', 'cappuccino,coffee,capp'],
        ['Coffee', 'Latte', '149', 'latte,coffee,cafe latte'],
        ['Coffee', 'Cold Coffee', '179', 'cold coffee,iced coffee,cold brew'],
        ['Coffee', 'Mocha', '169', 'mocha,chocolate coffee'],
        ['Tea', 'Masala Chai', '49', 'chai,masala tea,indian tea'],
        ['Tea', 'Green Tea', '79', 'green tea,greentea'],
        ['Tea', 'Ginger Tea', '59', 'ginger tea,adrak'],
        ['Burgers', 'Veg Burger', '149', 'veg burger,vegetarian burger'],
        ['Burgers', 'Cheese Burger', '179', 'cheese burger,cheese'],
        ['Burgers', 'Chicken Burger', '189', 'chicken burger,non-veg'],
        ['Burgers', 'Zinger Burger', '229', 'zinger,zinger burger'],
        ['Pizza', 'Margherita', '199', 'margherita,pizza,veg pizza'],
        ['Pizza', 'Farmhouse', '249', 'farmhouse,pizza,veg'],
        ['Pizza', 'Peppy Paneer', '279', 'peppy paneer,paneer pizza'],
        ['Pizza', 'Chicken Supreme', '349', 'chicken supreme,non-veg pizza'],
        ['Pasta', 'Red Sauce Pasta', '179', 'red sauce pasta,pasta,marinara'],
        ['Pasta', 'White Sauce Pasta', '199', 'white sauce pasta,alfredo,pasta'],
        ['Pasta', 'Pink Sauce Pasta', '219', 'pink sauce pasta,pasta'],
        ['Pasta', 'Chicken Pasta', '249', 'chicken pasta,non-veg pasta'],
        ['Desserts', 'Chocolate Brownie', '99', 'brownie,chocolate,dessert'],
        ['Desserts', 'Cheese Cake', '149', 'cheesecake,cake,cheese'],
        ['Desserts', 'Ice Cream', '79', 'icecream,ice cream,gelo'],
        ['Desserts', 'Chocolate Lava Cake', '179', 'lava cake,chocolate cake'],
        ['Beverages', 'Mango Smoothie', '129', 'mango,smoothie,shake'],
        ['Beverages', 'Oreo Shake', '149', 'oreo,shake,milkshake'],
        ['Beverages', 'Fresh Lime', '69', 'lime,lemon, nimbu'],
        ['Beverages', 'Mineral Water', '20', 'water,bottle,drinking'],
    ]
    
    # Add data
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num).value = value
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 30
    
    # Save
    wb.save("cafe_menu_template.xlsx")
    print("✅ Created: cafe_menu_template.xlsx")
    return "cafe_menu_template.xlsx"


def setup_cafe():
    """Interactive cafe setup"""
    print("\n" + "="*50)
    print("🏪 CAFE SETUP WIZARD")
    print("="*50 + "\n")
    
    name = input("Cafe Name: ").strip()
    address = input("Address: ").strip()
    phone = input("Phone Number: ").strip()
    
    return name, address, phone


def main():
    """Main function"""
    print("""
╔═══════════════════════════════════════════════════════════════╗
║                                                               ║
║       🏪 CAFE MENU TO AUTO-REPLY CONVERTER 🏪               ║
║                                                               ║
║       Convert your Excel menu to WhatsApp Bot keywords!       ║
║                                                               ║
╚═══════════════════════════════════════════════════════════════╝
    """)
    
    converter = CafeMenuConverter()
    
    print("\nOptions:")
    print("  [1] Create sample Excel menu template")
    print("  [2] Convert Excel menu to keywords")
    print("  [3] Convert JSON keywords file")
    print("  [4] Setup cafe info")
    print("  [5] Import to bot database")
    print("  [0] Exit\n")
    
    choice = input("Enter choice: ").strip()
    
    if choice == "1":
        create_sample_excel()
        
    elif choice == "2":
        excel_path = input("\nEnter Excel file path (e.g., my_menu.xlsx): ").strip()
        if os.path.exists(excel_path):
            keywords = converter.convert_from_excel(excel_path)
            converter.show_preview()
            
            # Save keywords
            output = input("\nSave keywords to file? (Enter filename or 'no'): ").strip()
            if output and output.lower() != 'no':
                converter.generate_keywords_file(output)
            
            # Import to bot
            import_now = input("\nImport to bot database? (y/n): ").strip().lower()
            if import_now == 'y':
                converter.import_to_bot()
        else:
            print(f"❌ File not found: {excel_path}")
    
    elif choice == "3":
        json_path = input("\nEnter JSON file path: ").strip()
        if os.path.exists(json_path):
            keywords = converter.convert_from_json(json_path)
            converter.keywords = keywords
            converter.show_preview()
        else:
            print(f"❌ File not found: {json_path}")
    
    elif choice == "4":
        name, address, phone = setup_cafe()
        converter.save_cafe_config(name, address, phone)
        print("\n✅ Cafe info saved!")
    
    elif choice == "5":
        print("\n📥 Importing to bot database...")
        converter.import_to_bot()
    
    elif choice == "0":
        print("\nGoodbye! 👋")
        sys.exit(0)
    
    else:
        print("\n❌ Invalid choice!")


if __name__ == "__main__":
    main()
